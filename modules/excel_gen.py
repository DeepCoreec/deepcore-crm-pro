"""DeepCore CRM Pro — Exportación a Excel."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

VERDE   = '10B981'
OSCURO  = '0c0c0e'
GRIS    = 'F2F4F8'
BLANCO  = 'FFFFFF'
AZUL    = '60A5FA'
AMARILLO= 'F59E0B'


def _borde():
    s = Side(style='thin', color='D1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _ajustar_columnas(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 4, 45)

def _titulo_hoja(ws, texto: str, empresa: str, ncols: int):
    rng = f'A1:{get_column_letter(ncols)}1'
    ws.merge_cells(rng)
    ws['A1'] = f'{empresa.upper()} — {texto}'
    ws['A1'].font      = Font(name='Calibri', bold=True, size=14, color=VERDE)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill      = _fill('0A1A10')

    rng2 = f'A2:{get_column_letter(ncols)}2'
    ws.merge_cells(rng2)
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font      = Font(name='Calibri', size=10, color='64748B')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16

def _headers(ws, row: int, headers: list, color=None):
    c = color or OSCURO
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.fill      = _fill(c)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = _borde()
    ws.row_dimensions[row].height = 20


def exportar_contactos_excel(contactos: list, config: dict, ruta: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contactos'
    empresa  = config.get('empresa', 'Mi Empresa')

    _titulo_hoja(ws, 'REPORTE DE CONTACTOS', empresa, 7)
    _headers(ws, 4, ['Nombre', 'Apellido', 'Empresa', 'Cargo', 'Email', 'Teléfono', 'Estado'])

    for i, c in enumerate(contactos, 5):
        fill = _fill(GRIS) if i % 2 == 0 else _fill(BLANCO)
        valores = [
            c.get('nombre',''), c.get('apellido',''),
            c.get('empresa_nombre') or '—', c.get('cargo') or '—',
            c.get('email') or '—', c.get('telefono') or '—',
            c.get('estado','')
        ]
        for j, val in enumerate(valores, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font      = Font(name='Calibri', size=10)
            cell.fill      = fill
            cell.border    = _borde()
            cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[i].height = 17

    _ajustar_columnas(ws)
    wb.save(ruta)
    return ruta


def exportar_pipeline_excel(oportunidades: list, config: dict, ruta: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pipeline'
    empresa = config.get('empresa', 'Mi Empresa')
    moneda  = config.get('moneda', '$')

    _titulo_hoja(ws, 'PIPELINE DE VENTAS', empresa, 7)
    _headers(ws, 4, ['Oportunidad', 'Contacto', 'Empresa', f'Valor ({moneda})', 'Etapa', 'Prob.', 'Cierre'], VERDE)

    total = 0.0
    for i, o in enumerate(oportunidades, 5):
        fill    = _fill(GRIS) if i % 2 == 0 else _fill(BLANCO)
        contacto = f"{o.get('contacto_nombre','') or ''} {o.get('contacto_apellido','') or ''}".strip() or '—'
        valor    = o.get('valor', 0) or 0
        total   += valor
        valores  = [
            o.get('nombre',''), contacto,
            o.get('empresa_nombre') or '—', valor,
            o.get('etapa',''), f"{o.get('probabilidad',0)}%",
            o.get('fecha_cierre') or '—'
        ]
        for j, val in enumerate(valores, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font      = Font(name='Calibri', size=10)
            cell.fill      = fill
            cell.border    = _borde()
            cell.alignment = Alignment(vertical='center')
            if j == 4:
                cell.number_format = f'"{moneda}"#,##0.00'
                cell.alignment     = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[i].height = 17

    last = len(oportunidades) + 5
    ws.cell(row=last, column=3, value='TOTAL PIPELINE').font = Font(bold=True, color=VERDE)
    tc = ws.cell(row=last, column=4, value=total)
    tc.font         = Font(bold=True, color=VERDE)
    tc.number_format = f'"{moneda}"#,##0.00'
    tc.alignment    = Alignment(horizontal='right')

    _ajustar_columnas(ws)
    wb.save(ruta)
    return ruta
